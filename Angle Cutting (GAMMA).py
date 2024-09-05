import math
from openpyxl import load_workbook
import os
import sys
import traceback

import shutil
from rich.console import Console
from rich.panel import Panel

console = Console()

# Print Ribbon
main_text = "ANGLE CUTTING SYSTEM - GAMMA"
small_text = "- Er. Shailesh Badhu"
panel_width = shutil.get_terminal_size().columns
panel = Panel(
    f"{main_text.center(panel_width)}\n{small_text.rjust(panel_width - 10)}",
    border_style="bold cyan"
)
console.print(panel)


# Add padded space after character
def pad_string(str, char):
    # Truncate the value from the right if its length is greater than char[key]
    if len(str) > char:
        str = str[:char]
    padded_str = str.ljust(char)
    return padded_str

def length_calculator(ldegree, rdegree, length, height):
    ltan_value = 1 / math.tan(math.radians(ldegree))
    rtan_value = 1 / math.tan(math.radians(rdegree))
    return round(length - abs(height * ltan_value + height * rtan_value),1)

def generate_lines(data, prefixes, char):
    try:
        calculated_len = "{:.10g}".format(length_calculator(data['left_tilt'], data['right_tilt'], data['len'], data['height']))
        if data['left_tilt'] < 0:
            data['ext_len'] = calculated_len
            data['int_len'] = data['len']
            del data['len']
        elif data['left_tilt'] > 0:
            data['ext_len'] = data['len']
            data['int_len'] = calculated_len
            del data['len']
        else:
            data['ext_len'] = data['len']
            data['int_len'] = data['len']
            del data['len']     
        padded_values = []


        # Pad each value in the data dictionary according to the character limits and append to the list
        for key in data:
            if key in prefixes:
                # Handle the keys with prefixes separately
                padded_values.append(prefixes[key] + pad_string(str(data[key]), char[key]))
            else:
                padded_values.append(pad_string(str(data[key]), char[key]))

        # Join all the padded values into a single string
        result = ''.join(padded_values) + "EXT\n"
        return result
    
    except Exception as e:
        log_error(e)
        return ""

def log_error(e):
    """Log error details to a file and print a user-friendly message."""
    with open("error_log.txt", "a") as log_file:
        log_file.write("An error occurred:\n")
        traceback.print_exc(file=log_file)
        log_file.write("\n")

    print("An error occurred. Please check 'error_log.txt' for details.")
    
def get_current_dir():
    """Get the directory where the script or executable is located."""
    if getattr(sys, 'frozen', False):
        # If the application is frozen (i.e., running as a .exe file)
        return os.path.dirname(sys.executable)
    else:
        # If the script is running normally (i.e., .py file)
        return os.path.dirname(os.path.abspath(__file__))
    

def main():
    try:
        char = {
            'barcode': 13,
            'code': 15,
            'ext_len': 7,
            'int_len': 7,
            'height': 5,
            'qty_char': 3,
            'left_tilt': 5,
            'right_tilt': 5,
            'Nr_piece': 5,
            'lot': 9,
            'trolley': 2,
            'order': 10,
            'frame': 3,
            'shutter': 5,
            'description': 100
        }

        prefixes = {
            'code': "C",
            'lot': "Lot",
            'trolley': "Casier",
            'order': "Cde",
            'frame': "Ch.",
            'shutter': "Ouv"
        }

        file_name = input("Save file as: ")
        file_name = file_name + ".txt"

        # Use the get_current_dir function to get the directory of the script/executable
        current_dir = get_current_dir()
        excel_file = os.path.join(current_dir, 'CUTTING_LIST_FOR_GAMMA.xlsx')

        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"Excel file not found: {excel_file}")

        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook['LIST']

        last_row = sheet.max_row
        for row in range(last_row, 0, -1):
            cell = sheet.cell(row=row, column=3)
            if cell.value is not None:
                last_value_row = row
                break


        # Define the row to check
        row_number = 9

        # Find the last column with data in the specified row
        last_column = 0
        for cell in sheet[row_number]:
            if cell.value is not None:
                last_column = max(last_column, cell.column)

        no_of_profiles = int((last_column - 3) / 2)
        to_print_des = [1]
        to_print_code = []
        lines = []
        for col in range(4, no_of_profiles * 2 + 4, 2):
            for row in range(9, last_value_row + 2, 1):
                if sheet.cell(row=row, column = col + 1).value is not None:
                    data = {
                        'barcode': "",
                        'code': str(sheet.cell(row=4, column=col).value),
                        'len': int(sheet.cell(row=row, column=col + 1).value),
                        'ext_len': 0,
                        'int_len': 0,
                        'height': int(sheet.cell(row=5, column=col).value),
                        'qty_char': int(sheet.cell(row=row, column=col).value),
                        'left_tilt': int(sheet.cell(row=6, column=col).value),
                        'right_tilt': int(sheet.cell(row=6, column=col + 1).value),
                        'Nr_piece': "",
                        'lot': "",
                        'trolley': "",
                        'order': str(sheet.cell(row=2, column=3).value),
                        'frame': "",
                        'shutter': "",
                        'description': str(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value is not None else None
                    }
                    if data['description'] is None:
                        data['description'] = str(sheet.cell(row=row-1,column=3).value)
                    if col == 4 and data['description'] != to_print_des[-1]:
                        if to_print_des == [1]:
                            to_print_des = []
                        to_print_des.append(data['description'])
                    line = generate_lines(data, prefixes, char)
                    lines.append(line)
            to_print_code.append(data['code'])
            
        lines_str = ''.join(lines)

        header = "\t  " + "\t".join([str(i + 1) for i in range(len(to_print_code))])
        print("\n\n" + header)

        # Print the to_print_code values in the next row
        print("\t  " + "\t".join(to_print_code))

        num_entries = len(to_print_des)

        # Print each string vertically with corresponding row numbers
        for i, des in enumerate(to_print_des, start=1):
            print(f"{i}  {des}")
            
        with open(file_name, "w") as file:
            file.write(lines_str)   
        print(f"\n\nFile '{file_name}' has been saved successfully.\n")

    except FileNotFoundError as e:
        log_error(e)
        print(f"File not found: {e.filename}")
    except ValueError as e:
        log_error(e)
        print(f"Value error: {str(e)}")
    except Exception as e:
        log_error(e)
        print("An unexpected error occurred.")
    finally:
        input("Press Enter to exit")

if __name__ == "__main__":
    main()
